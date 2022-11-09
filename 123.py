'''
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

wbd= load_workbook('C:/GitHub/ITC Upgrade Test/ISO_codes.xlsx')
wsd = wbd.active

partArr = []

for code in range(252):
    partArr.append(str(wsd['C'+str(2+code)].value))

print(partArr)
'''
from openpyxl import Workbook
wb = Workbook(write_only=True)
ws = wb.create_sheet()

# now we'll fill it with 100 rows x 200 columns
for irow in range(100):
    ws.append(['1', ,'3'])

# save the file
wb.save('new_big_file.xlsx') 
