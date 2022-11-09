import sys, os
import win32com.client
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

#create a result file
result_path='C:/GitHub/ITC Upgrade Test/Tables Acces (output)/Israel import.xlsx'

#потом надо будет чтоб формат брался от path
format = 'import'

word_result = load_workbook(result_path)
#wrs - word result sheet 
wrs = word_result.active

def create_header(result_table, reporter):
    #create header потом допилить для экспорта
    if (format == 'import'):
        wrs['A1'] = 'Reporter'
        wrs['B1'] = 'Partner'
        wrs['C1'] = 'Product code'
        wrs['D1'] = 'Product label'
        wrs['E1'] = reporter+' imp from partn, $1000, 2002'
        wrs['F1'] = reporter+' imp from partn, $1000, 2003'
        wrs['G1'] = reporter+' imp from partn, $1000, 2004'
        wrs['H1'] = reporter+' imp from partn, $1000, 2005'
        wrs['I1'] = reporter+' imp from partn, $1000, 2006'
        wrs['J1'] = reporter+' imp from partn, $1000, 2007'
        wrs['K1'] = reporter+' imp from partn, $1000, 2008'
        wrs['L1'] = reporter+' imp from partn, $1000, 2009'
        wrs['M1'] = reporter+' imp from partn, $1000, 2010'
        wrs['N1'] = reporter+' imp from partn, $1000, 2011'
        wrs['O1'] = reporter+' imp from partn, $1000, 2012'
        wrs['P1'] = reporter+' imp from partn, $1000, 2013'
        wrs['Q1'] = reporter+' imp from partn, $1000, 2014'
        wrs['R1'] = reporter+' imp from partn, $1000, 2015'
        wrs['S1'] = reporter+' imp from partn, $1000, 2016'
        wrs['T1'] = reporter+' imp from partn, $1000, 2017'
        wrs['U1'] = reporter+' imp from partn, $1000, 2018'
        wrs['V1'] = reporter+' imp from partn, $1000, 2019'
        wrs['W1'] = reporter+' imp from partn, $1000, 2020'
        wrs['X1'] = reporter+' imp from partn, $1000, 2021'

        wrs['Y1'] = 'Partner exp to Wrld, $1000, 2002'
        wrs['Z1'] = 'Partner exp to Wrld, $1000, 2003'
        wrs['AA1'] = 'Partner exp to Wrld, $1000, 2004'
        wrs['AB1'] = 'Partner exp to Wrld, $1000, 2005'
        wrs['AC1'] = 'Partner exp to Wrld, $1000, 2006'
        wrs['AD1'] = 'Partner exp to Wrld, $1000, 2007'
        wrs['AE1'] = 'Partner exp to Wrld, $1000, 2008'
        wrs['AF1'] = 'Partner exp to Wrld, $1000, 2009'
        wrs['AG1'] = 'Partner exp to Wrld, $1000, 2010'
        wrs['AH1'] = 'Partner exp to Wrld, $1000, 2011'
        wrs['AI1'] = 'Partner exp to Wrld, $1000, 2012'
        wrs['AJ1'] = 'Partner exp to Wrld, $1000, 2013'
        wrs['AK1'] = 'Partner exp to Wrld, $1000, 2014'
        wrs['AL1'] = 'Partner exp to Wrld, $1000, 2015'
        wrs['AM1'] = 'Partner exp to Wrld, $1000, 2016'
        wrs['AN1'] = 'Partner exp to Wrld, $1000, 2017'
        wrs['AO1'] = 'Partner exp to Wrld, $1000, 2018'
        wrs['AP1'] = 'Partner exp to Wrld, $1000, 2019'
        wrs['AQ1'] = 'Partner exp to Wrld, $1000, 2020'
        wrs['AR1'] = 'Partner exp to Wrld, $1000, 2021'

        wrs['AS1'] = reporter+' imp from Wrld, $1000, 2002'
        wrs['AT1'] = reporter+' imp from Wrld, $1000, 2003'
        wrs['AU1'] = reporter+' imp from Wrld, $1000, 2004'
        wrs['AV1'] = reporter+' imp from Wrld, $1000, 2005'
        wrs['AW1'] = reporter+' imp from Wrld, $1000, 2006'
        wrs['AX1'] = reporter+' imp from Wrld, $1000, 2007'
        wrs['AY1'] = reporter+' imp from Wrld, $1000, 2008'
        wrs['AZ1'] = reporter+' imp from Wrld, $1000, 2009'
        wrs['BA1'] = reporter+' imp from Wrld, $1000, 2010'
        wrs['BB1'] = reporter+' imp from Wrld, $1000, 2011'
        wrs['BC1'] = reporter+' imp from Wrld, $1000, 2012'
        wrs['BD1'] = reporter+' imp from Wrld, $1000, 2013'
        wrs['BE1'] = reporter+' imp from Wrld, $1000, 2014'
        wrs['BF1'] = reporter+' imp from Wrld, $1000, 2015'
        wrs['BG1'] = reporter+' imp from Wrld, $1000, 2016'
        wrs['BH1'] = reporter+' imp from Wrld, $1000, 2017'
        wrs['BI1'] = reporter+' imp from Wrld, $1000, 2018'
        wrs['BJ1'] = reporter+' imp from Wrld, $1000, 2019'
        wrs['BK1'] = reporter+' imp from Wrld, $1000, 2020'
        wrs['BL1'] = reporter+' imp from Wrld, $1000, 2021'

    #header format
    wrs.row_dimensions[1].height = 80
    for cellq in range(70):
        cellq = cellq + 1
        wrs.cell(1, cellq).alignment = Alignment(wrap_text=True,vertical='top') 

    return wrs


directory = 'C:\\GitHub\\ITC Upgrade Test\\Tables 25\\Imports\\Israel\\'

#last row position
last_raw = 0
q = False #for 1 update RAM

#for all file in directory MERGE
for file in os.listdir(directory):
    print('Ф ',file)
    #load 
    wbd = load_workbook(directory+str(file))
    wsd = wbd.active

    #Find count of product
    productCount = 0
    for a in range(7000):
        a = a + 1
        if (wsd['A'+str(a)].value == None and a > 13):
            productCount = a - 14
            break

    #find reporter and partner
    mainStr = str(wsd['A1'].value[:len(wsd['A1'].value)-1])
    mainStr = mainStr[mainStr.find('and')+28:]

    reporter = mainStr[:mainStr.find('and')-1]
    partner = mainStr[mainStr.find('and')+4:]

    #If we going first time - create header
    if last_raw == 0:
        wrs = create_header(wrs, reporter)

    #create 2 columns
    for row in range(productCount):
        row = row+2+last_raw
        wrs['A'+str(row)] = reporter
        wrs['B'+str(row)] = partner

    #product_codes and names
    for product in range(productCount):
        wrs['C' + str(product + 2+last_raw)] = wsd['A'+str(14+product)].value[1:]
        wrs['D' + str(product + 2+last_raw)] = wsd['B'+str(14+product)].value

    for row in range(productCount):
        for column1 in range(20):
            wrs.cell(row=row+2+last_raw, column=5+column1).value = wsd.cell(row=row+14, column=3+column1).value
        for column2 in range(20):
            wrs.cell(row=row+2+last_raw, column=25+column2).value = wsd.cell(row=row+14, column=24+column2).value
        for column2 in range(20):
            wrs.cell(row=row+2+last_raw, column=45+column2).value = wsd.cell(row=row+14, column=45+column2).value
    
    last_raw = last_raw + productCount
    print(last_raw)

word_result.save(result_path)